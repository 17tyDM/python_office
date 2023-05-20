from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

ws2 = wb.create_sheet('practice')
num = 0
for i in range(1,11):
    for j in ['A','B','C','D','E','F','G','H','I','J']:
        num += 1
        ws2[f"{j}{i}"] = num
wb.save("sample.xlsx")